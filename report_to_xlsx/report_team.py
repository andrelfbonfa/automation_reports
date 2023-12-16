import cx_Oracle
import openpyxl
import config_send_email
import time
from datetime import date ,datetime, timezone

now = datetime.now() # current date and time
print(now)

data_atual = now.strftime("%d-%m-20%y")
print (data_atual)

# Conectar ao banco de dados Oracle
connection = cx_Oracle.connect('user/pass@host/db',encoding='UTF-8', nencoding='UTF-8' )

# Executar a consulta
cursor = connection.cursor()
with open("sql/query.sql") as f:
     plsql = f.read()
     cursor.execute(plsql)

# Obter os resultados da consulta
     results = cursor.fetchall()

# Criar ou abrir o arquivo XLSX
     wb = openpyxl.Workbook()

# Adicionar uma aba
     sheet = wb.active
     sheet.title = 'report'

# Escrever os cabeçalhos na primeira linha
     for i, column_title in enumerate(cursor.description):
         sheet.cell(row=1, column=i + 1).value = column_title[0]

# Escrever os resultados da consulta na planilha
     for i, row in enumerate(results):
         for j, value in enumerate(row):
             sheet.cell(row=i + 2, column=j + 1).value = value




   
# Salvar o arquivo XLSX
wb.save("report_name-"+data_atual+".xlsx")

config_send_email.enviaemail("Assunto - "+data_atual+"", "Olá, Segue em anexo a planilha atualizada.", "report_name-"+data_atual+".xlsx")

# Fechar a conexão com o banco de dados Oracle
cursor.close()
connection.close()